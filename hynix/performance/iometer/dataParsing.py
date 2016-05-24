'''
Created on 2016. 3. 15.

@author: lee chanjoo
'''

import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import csv, itertools, cx_Oracle,datetime, os, re , math , shutil
from _sqlite3 import Row
from itertools import product
from telnetlib import theNULL
from test.test_datetime import DSTEND

def insertQuery(result):
    insert = '''INSERT INTO PERFORMANCE 
                    (FOLDER_NAME, 
                    TOOL,
                    DATA_SRC, 
                    SPEC, 
                    SPEC_LABEL, 
                    QUEUE_DEPTH, 
                    FIELD, 
                    MEASURE)
                VALUES 
                    (:1, :2, :3, :4, :5, :6, :7, :8)'''
 
    try:
        cur.prepare(insert)
        cur.executemany(None, result)
#         con.commit()
         
    except:
        pass
 
 
def data_source(path):
    data_src = os.path.basename(path)
    return data_src
 
 
def lists(folder_name, tool, data_src, _spec, _field, _measure):
    result = [(folder_name, 
               tool,     
               data_src, 
               _spec,    
               '',       
               '1',      
               _field,   
               _measure)]
 
    insertQuery(result)
    print result
          
  
def parseXml(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
      
    tree = ET.parse(file_name)
    root = tree.getroot()
    tag_text = root.findtext(tag_name)
    tag_val = tag_text.split(' ')[0]
    value = tag_name.split('/')  
     
    if value[0] == 'SeqTest':
        lists(folder_name, tool, data_src, _spec=value[0], _field=value[1], _measure=tag_val)
    elif value[0] == 'Random4K1TTest':
        lists(folder_name, tool, data_src, _spec=value[0], _field=value[1], _measure=tag_val)
    elif value[0] == 'Random4K64TTest':
        lists(folder_name, tool, data_src, _spec=value[0], _field=value[1], _measure=tag_val)
         
 
def parseAccTime(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
         
    tree = ET.parse(file_name)
    root = tree.getroot()
    tag_text = root.findtext(tag_name)
    tag_val = tag_text.split(' ')[0]
    _float = float(tag_val)
    _float3 = '{0:.3f}'.format(_float)
    value = tag_name.split('/')
     
    lists(folder_name, tool, data_src, _spec=value[0], _field='-', _measure=_float3)
 
 
def parseMark8Score(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
     
    tree = ET.parse(file_name)
    root = tree.getroot()
    tag_text = root.findtext(tag_name)
    tag_val = tag_text.split(' ')[0]
    value = tag_name.split('/')
     
    lists(folder_name, tool, data_src, _spec=value[2], _field='-', _measure=tag_val)
     
     
def parseMark8(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
     
    tree = ET.parse(file_name)
    root = tree.getroot()
    tag_text = root.findtext(tag_name)
    tag_val = tag_text.split(' ')[0]
    val = int(tag_val)/1024/1024
    value = tag_name.split('/')
     
    lists(folder_name, tool, data_src, _spec=value[2], _field='-', _measure=val)
     
 
def parseFdr(folder_name, tool, file, overall_tag, scenario_tag, overallperformance_tag):
     
    data_src = data_source(file)
         
    tree = ET.parse(file)
    root = tree.getroot()
    tag_text = root.findtext(overallperformance_tag)
     
    for overall in root.iter(overall_tag):
        for scenario in overall.iter(scenario_tag):
            lists(folder_name, tool, data_src, _spec=scenario.attrib['name'], _field='-', _measure=scenario.attrib['performance'])   
    lists(folder_name, tool, data_src, _spec='overallperformance', _field='-', _measure=tag_text)
     
 
def parseTxt(folder_name, tool, file, line1, line2, spec, p):
     
    data_src = data_source(file)        
         
    with open(file) as f:
        for i, j in enumerate(f):
            pass
     
    num = i + 1
         
    with open(file) as f:        
        for line in itertools.islice(f, num-line1, num-line2):
            if p == pattern_txt: 
                match = p.search(line)
                if match:
                    data = match.groupdict()['flt']
                    raw_data = '{0:.2f}'.format(float(data))
                    lists(folder_name, tool, data_src, _spec=spec, _field='-', _measure=raw_data)
     
            elif p == pattern_score:
                match = p.search(line)
                if match:
                    score_data = match.groupdict()['score']
                    lists(folder_name, tool, data_src, _spec=spec, _field='-', _measure=score_data)
 
 
def parseTxtPor(folder_name, tool, file, pattern, spec):
     
    data_src = data_source(file)
     
    with open(file) as f:
        for line in f:
            match = pattern.search(line)
            if match:
                data = match.groupdict()['por']
                value = int(data)/1000
                lists(folder_name, tool, data_src, _spec=spec, _field='tBOOTUP', _measure=value)
                 
             
def parseXlsx(folder_name, tool, file, pattern, race=file):
     
    data_src = data_source(file)
     
    wb = load_workbook(filename=file)
       
    if race == CDM301:
        sheet_ranges = wb['Sheet1']
    elif race == CDM503:
        sheet_ranges = wb['Sheet1']
    elif race == PCMARK_VANTAGE:
        sheet_ranges = wb['Result 1']
     
    pcmark_lists = [140, 142, 144, 146, 148, 150, 152, 154, 139]
     
    if race == CDM301:
        iCDM301 = 0
        for row in sheet_ranges.iter_rows():
            for cell in row:
                if cell.value is not None:
                    match =  pattern.search(cell.value)    
                    if match:
                        iCDM301 = iCDM301 + 1
                        if iCDM301 == 1:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Read', _measure=match.groupdict()['num'])
                        elif iCDM301 == 2:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Write', _measure=match.groupdict()['num'])
                        elif iCDM301 == 3:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 512KB', _measure=match.groupdict()['num'])
                        elif iCDM301 == 4:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 512KB', _measure=match.groupdict()['num'])
                        elif iCDM301 == 5:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 4KB (QD=1)', _measure=match.groupdict()['num'])
                        elif iCDM301 == 6:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 4KB (QD=1)', _measure=match.groupdict()['num'])
                        elif iCDM301 == 7:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 4KB (QD=32)', _measure=match.groupdict()['num'])
                        elif iCDM301 == 8:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 4KB (QD=32)', _measure=match.groupdict()['num'])                    
               
    elif race == CDM503:
        iCDM503 = 0
        for row in sheet_ranges.iter_rows():  
            for cell in row:
                if cell.value is not None:
                    match =  pattern.search(cell.value)      
                    if match:
                        iCDM503 = iCDM503 + 1
                        if iCDM503 == 1:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Read (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 2:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Write (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 3:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 4:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 5:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='REad (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 6:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Write (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 7:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num'])   
                        elif iCDM503 == 8:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num']) 
 
    elif race == PCMARK_VANTAGE:
        for i in pcmark_lists:
            if i == 140:
                lists(folder_name, tool, data_src, _spec='HDD1', _field='Windows Defender', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 142:
                lists(folder_name, tool, data_src, _spec='HDD2', _field='Gaming', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 144:
                lists(folder_name, tool, data_src, _spec='HDD3', _field='Importing pictures to Windows Photo Gallery', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 146:
                lists(folder_name, tool, data_src, _spec='HDD4', _field='Windows Vista startup', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 148:
                lists(folder_name, tool, data_src, _spec='HDD5', _field='Video editing using Windows Movie Maker', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 150:
                lists(folder_name, tool, data_src, _spec='HDD6', _field='Windows Media Center', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 152:
                lists(folder_name, tool, data_src, _spec='HDD7', _field='Adding music to Windows Media Player', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 154:
                lists(folder_name, tool, data_src, _spec='HDD8', _field='Application loading', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 139:
                lists(folder_name, tool, data_src, _spec='Score', _field='-', _measure=math.trunc(sheet_ranges['C' + str(i)].value))



#Iometer start
matrix = []
vToday = datetime.datetime.now().strftime("%Y%m%d")
query = ''
path = "//10.15.202.120/das_input/Solution/SSD_Performance"
dirsIometer = os.listdir(path);
# folderHeaders = ['SYSTEM', 'OPERATING_SYSTEM', 'VENDOR', 'PRODUCT_NAME', 'NAND_TECH', 'CELL_TYPE','CONTROLLER', 'FORM_FACTOR', 'CAPACITY', 'FIRMWARE', 'SLC_BUFFER', 'MEASURE_DT']
folderHeaders = ['MEASURE_DT','VENDOR','PRODUCT_NAME', 'CONTROLLER', 'NAND_TECH','CELL_TYPE', 'FORM_FACTOR','CAPACITY','FIRMWARE','SLC_BUFFER','SERIAL_NUMBER','TEST_COUNT']
gFolderInfos = []
filesIometer = []
for subdir in dirsIometer:
    if subdir == 'parsed':
        continue
    pathIometer = path + '/' + subdir + '/Iometer'
    dirIometer = os.listdir(pathIometer)
    
    #folderInfos
    gFolderInfos.append({"folderName":subdir,"data":dict(zip(folderHeaders, subdir.split("_")))});
    
    for file in dirIometer:
        fileObj = {"filePath":pathIometer + '/' + file,"folderName":subdir,"fileName":file}
        filesIometer.append(fileObj)

# 01.Max Throughput.csv
for file in filesIometer:
    f = open(file['filePath'], 'r')
     
#      d_reader = csv.DictReader(f)
#      headers = d_reader.fieldnames
#      for line in d_reader:
#          matrix.append(line)
     
    #folderInfo
    folderInfos = file['folderName'].split("_")
    dicFolder = dict(zip(folderHeaders, folderInfos))
    csvReader = csv.reader(f)
    header = []
     
    for row in csvReader:
        if row[0] == "'Target Type" and len(header) == 0 :
            header = row
        if row[0] ==  "WORKER" :
            dictionary = dict(zip(header, row))
            if dictionary["Access Specification Name"] == 'Precondition':
                continue
            # delete ' charater
            dictionary["Target Type"] = dictionary["'Target Type"]
            del dictionary["'Target Type"]
            dictionary["meta"] = {}
            dictionary["meta"]["fileName"] = file['fileName'] #"Max Throughput.csv"
            dictionary["meta"]["folderName"] = file['folderName']
            dictionary["meta"]["tool"] = 'Iometer'
             
            #header infos
            for head in folderHeaders:
                dictionary["meta"][head] = dicFolder[head]
             
            dictionary["keys"] = {}
            dictionary["keys"]["Access Specification Name"] = dictionary["Access Specification Name"]
            del dictionary["Access Specification Name"]
            dictionary["keys"]["Queue Depth"] = dictionary["Queue Depth"]
            del dictionary["Queue Depth"]
             
            matrix.append(dictionary)
#             print dictionary
       
    f.close()
     
# print len(matrix)
try:
     
    vProduct = '' 
    con_str = 'swdashboard/swdashboard@166.125.19.98:1521/APS'
    con = cx_Oracle.connect(con_str)
    cur = con.cursor()
     
    for folders in gFolderInfos:   
        i = 0
         
        query = "delete from PERFORMANCE_FOLDER where " 
        query += " FOLDER_NAME ='" + folders['folderName'] +"'"
        cur.execute(query)
         
        query = "delete from PERFORMANCE where " 
        query += " FOLDER_NAME ='" + folders['folderName'] +"'"
#         query += " and tool = 'Iometer'"
        cur.execute(query)
         
        query = "insert into PERFORMANCE_FOLDER(FOLDER_NAME, MEASURE_DT, VENDOR, PRODUCT_NAME, CONTROLLER, NAND_TECH, CELL_TYPE, FORM_FACTOR, CAPACITY, FIRMWARE, SLC_BUFFER, SERIAL_NUMBER, TEST_COUNT)"
        query += " values("
        query += "'" + folders['folderName'] + "'"
        for head in folderHeaders:
            query += ",'" + folders['data'][head] + "'" 
            i = i + 1
             
        query += ")"
        cur.execute(query)
    con.commit()   
         
         
    i = 0
    for m in matrix:
        i = i+1
        print repr(i) + '(' + repr(len(matrix)) + ')'
        for mKey in m: 
            if type(m[mKey]) is dict:
                continue
            query = " insert into PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC, SPEC_LABEL, QUEUE_DEPTH, FIELD, MEASURE)"
            query += " values('" + m["meta"]['folderName'] +"','" + m["meta"]['tool'] +"','" + m["meta"]['fileName'] +"', '" + m["keys"]['Access Specification Name'] +"','', '" + m["keys"]["Queue Depth"] +"', '"+ mKey +"',  '" + m[mKey] + "'  )"
             
            cur.execute(query)
    con.commit()
    #Iometer End
    #not Iometer
    # Tag
    seq_test_r_tag = 'SeqTest/Read'
    seq_test_w_tag = 'SeqTest/Write'
    Random4K1TTest_r_tag = 'Random4K1TTest/Read'
    Random4K1TTest_w_tag = 'Random4K1TTest/Write'        
    Random4K64TTest_r_tag = 'Random4K64TTest/Read'
    Random4K64TTest_w_tag = 'Random4K64TTest/Write'     
    acc_time_r_tag = 'AccTimeTest/Read'
    acc_time_w_tag = 'AccTimeTest/Write'
    overall_tag = '{http://www.bapco.com}overall'
    scenario_tag = '{http://www.bapco.com}scenario' 
    overallperformance_tag = '{http://www.bapco.com}data/{http://www.bapco.com}overallperformance'
    storage_score = 'results/result/storage_score'
    storage_bandwidth = 'results/result/storage_bandwidth'
     
    # Regex
    pattern = re.compile(r'((?P<num>\d*).\d+)\sMB\/s')
    pattern_txt = re.compile(r'((?P<flt>\d*.\d+))\sMB\/s')
    pattern_score = re.compile(r'(score:\s(?P<score>\d+)\s)')
    pattern_tBOOTUP = re.compile(r'tBOOTUP\s:\s(?P<por>\d+)')
                 
    pwd = '//10.15.202.120/das_input/Solution/SSD_Performance'
 
    for path, dirs, files in os.walk(pwd):
        dirs[:] = [dir for dir in dirs if dir != "Iometer" and dir not in("parsed") ]
        for file in files:
            if os.path.basename(file) != 'readme.txt':
                current_dir_path = os.path.join(path, file)             
                path_sep = os.path.sep
                components = current_dir_path.split(path_sep)                    
                 
                if file == 'PCMARK_VANTAGE.xlsx':
                    PCMARK_VANTAGE = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, PCMARK_VANTAGE, pattern, race=PCMARK_VANTAGE)
                     
                elif file == 'CDM301.xlsx':
                    CDM301 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]        
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, CDM301, pattern, race=CDM301)   
                     
                elif file == 'CDM503.xlsx':
                    CDM503 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, CDM503, pattern, race=CDM503)
                     
                elif file == 'AS-SSD1.6.xml':
                    AS_SSDv16 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]                        
                    print '----AS-SSDv1.6'
                    parseXml(folder_name, tool, AS_SSDv16, seq_test_r_tag)
                    parseXml(folder_name, tool, AS_SSDv16, seq_test_w_tag)
                    parseXml(folder_name, tool, AS_SSDv16, Random4K1TTest_r_tag)
                    parseXml(folder_name, tool, AS_SSDv16, Random4K1TTest_w_tag)
                    parseXml(folder_name, tool, AS_SSDv16, Random4K64TTest_r_tag)
                    parseXml(folder_name, tool, AS_SSDv16, Random4K64TTest_w_tag)
                    parseAccTime(folder_name, tool, AS_SSDv16, acc_time_r_tag)
                    parseAccTime(folder_name, tool, AS_SSDv16, acc_time_w_tag)
                     
                elif file == 'AS-SSD1.8.xml':
                    AS_SSDv18 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----AS-SSDv1.8'
                    parseXml(folder_name, tool, AS_SSDv18, seq_test_r_tag)
                    parseXml(folder_name, tool, AS_SSDv18, seq_test_w_tag)           
                    parseXml(folder_name, tool, AS_SSDv18, Random4K1TTest_r_tag)
                    parseXml(folder_name, tool, AS_SSDv18, Random4K1TTest_w_tag)
                    parseXml(folder_name, tool, AS_SSDv18, Random4K64TTest_r_tag)
                    parseXml(folder_name, tool, AS_SSDv18, Random4K64TTest_w_tag)
                    parseAccTime(folder_name, tool, AS_SSDv18, acc_time_r_tag)
                    parseAccTime(folder_name, tool, AS_SSDv18, acc_time_w_tag)
                     
                elif file == 'PCMARK7.txt':
                    PCMARK7 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseTxt'
                    parseTxt(folder_name, tool, PCMARK7, 21, 20, 'Windows Defender', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 20, 19, 'Importing pictures', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 19, 18, 'Video editing', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 18, 17, 'Windows Media Center', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 17, 16, 'Adding Music', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 16, 15, 'Starting applications', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 15, 14, 'Gaming', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 22, 21, 'Secondary storage score', p=pattern_score)
                     
                elif file == 'PCMARK8.xml':
                    PCMARK8 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----PCMARK8'
                    parseMark8Score(folder_name, tool, PCMARK8, storage_score)
                    parseMark8(folder_name, tool, PCMARK8, storage_bandwidth)
                     
                elif file == 'SYSMARK2014.fdr':
                    SYSMARK2014 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseFdr'
                    parseFdr(folder_name, tool, SYSMARK2014, overall_tag, scenario_tag, overallperformance_tag)
                     
                elif file == 'POR.txt':
                    POR = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]                        
                    print '----parseTxtPor/Spor'
                    parseTxtPor(folder_name, tool, POR, pattern_tBOOTUP, spec='POR time')
                     
                elif file == 'SPOR.txt':
                    SPOR = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseTxtPor/Spor'
                    parseTxtPor(folder_name, tool, SPOR, pattern_tBOOTUP, spec='SPOR time')
         
    con.commit()
    # file move to parsed
    for folders in gFolderInfos:
        src = pwd + '/' + folders['folderName']
        dst = pwd + '/parsed/' + folders['folderName']
        if os.path.exists(dst):   
            shutil.rmtree(dst)
        shutil.copytree(src,dst)
        shutil.rmtree(src)
    
except cx_Oracle.DatabaseError ,ex:
    error, = ex.args    
    print query
    print 'Error Inserting Field Base'
    print 'Error.code    =', error.code
    print 'Error.message =', error.message
    print 'Error.offset  =', error.offset
    con.rollback()
    
finally :
    if cur:
        cur.close()     
    if con:
        con.close()    