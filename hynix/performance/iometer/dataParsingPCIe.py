'''
Created on 2016. 3. 15.

@author: lee chanjoo
'''

import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import csv, itertools, cx_Oracle,datetime, os, re , math , shutil
from _sqlite3 import Row
from itertools import product
from telnetlib import theNULL
from test.test_datetime import DSTEND
import subprocess
from pip._vendor.pkg_resources import null_ns_handler
from lxml import etree
import urllib

def insertQuery(result):
    insert = '''INSERT INTO PERFORMANCE 
                    (FOLDER_NAME, 
                    TOOL,
                    DATA_SRC, 
                    SPEC, 
                    SPEC_LABEL, 
                    QUEUE_DEPTH, 
                    FIELD, 
                    MEASURE)
                VALUES 
                    (:1, :2, :3, :4, :5, :6, :7, :8)'''
 
    try:
        cur.prepare(insert)
        cur.executemany(None, result)
#         con.commit()
         
    except:
        pass
 
     
def data_source(path):
    data_src = os.path.basename(path)
    return data_src
 
 
def lists(folder_name, tool, data_src, _spec, _field, _measure):
    result = [(folder_name, 
               tool,     
               data_src, 
               _spec,    
               '',       
               '1',      
               _field,   
               _measure)]
 
    insertQuery(result)
    print result
    
# PCIe cdm 
def lists1(folder_name, tool, data_src, _spec,_spec1, _field, _measure):
    result = [(folder_name, 
               tool,     
               data_src, 
               _spec,   
               _spec1, 
               '',       
               '1',      
               _field,   
               _measure)]
 
    insertQuery1(result)
    print result 
# PCIe cdm 
def insertQuery1(result):
    insert = '''INSERT INTO PERFORMANCE 
                    (FOLDER_NAME, 
                    TOOL,
                    DATA_SRC, 
                    SPEC, 
                    SPEC1,
                    SPEC_LABEL, 
                    QUEUE_DEPTH, 
                    FIELD, 
                    MEASURE)
                VALUES 
                    (:1, :2, :3, :4, :5, :6, :7, :8, :9)'''
 
    try:
        cur.prepare(insert)
        cur.executemany(None, result)
#         con.commit()
         
    except:
        pass
		
		
        
def parseXml(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
      
    tree = ET.parse(file_name)
    root = tree.getroot()
    tag_text = root.findtext(tag_name)
    tag_val = tag_text.split(' ')[0]
    value = tag_name.split('/')  
     
    if value[0] == 'SeqTest':
        lists(folder_name, tool, data_src, _spec=value[0], _field=value[1], _measure=tag_val)
    elif value[0] == 'Random4K1TTest':
        lists(folder_name, tool, data_src, _spec=value[0], _field=value[1], _measure=tag_val)
    elif value[0] == 'Random4K64TTest':
        lists(folder_name, tool, data_src, _spec=value[0], _field=value[1], _measure=tag_val)

def parseXmlGneric(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
      
    tree = ET.parse(file_name)
    root = tree.getroot()
    vList = tree.find(tag_name)
    vElements = vList.getchildren()
    
    for el in vElements:        
        tag_text = el.tag
        tag_val = el.text
    
        lists(folder_name, tool, data_src, _spec=tag_text, _field=tag_text, _measure=tag_val)
       		
 
def parseAccTime(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
         
    tree = ET.parse(file_name)
    root = tree.getroot()
    tag_text = root.findtext(tag_name)
    tag_val = tag_text.split(' ')[0]
    _float = float(tag_val)
    _float3 = '{0:.3f}'.format(_float)
    value = tag_name.split('/')
     
    lists(folder_name, tool, data_src, _spec=value[0], _field='-', _measure=_float3)
 
 
def parseMark8Score(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
     
    tree = ET.parse(file_name)
    root = tree.getroot()
    tag_text = root.findtext(tag_name)
    tag_val = tag_text.split(' ')[0]
    value = tag_name.split('/')
     
    lists(folder_name, tool, data_src, _spec=value[2], _field='-', _measure=tag_val)
     
     
def parseMark8(folder_name, tool, file_name, tag_name):
     
    data_src = data_source(file_name)
     
    tree = ET.parse(file_name)
    root = tree.getroot()
    tag_text = root.findtext(tag_name)
    tag_val = tag_text.split(' ')[0]
    val = int(tag_val)/1024/1024
    value = tag_name.split('/')
     
    lists(folder_name, tool, data_src, _spec=value[2], _field='-', _measure=val)
     
	 
def parseFdr(folder_name, tool, file, overall_tag, scenario_tag, overallperformance_tag):
     
    data_src = data_source(file)
         
    tree = ET.parse(file)
    root = tree.getroot()
    tag_text = root.findtext(overallperformance_tag)
     
    for overall in root.iter(overall_tag):
        for scenario in overall.iter(scenario_tag):
            lists(folder_name, tool, data_src, _spec=scenario.attrib['name'], _field='-', _measure=scenario.attrib['performance'])   
    lists(folder_name, tool, data_src, _spec='overallperformance', _field='-', _measure=tag_text)
     
 
def parseTxt(folder_name, tool, file, line1, line2, spec, p):
     
    data_src = data_source(file)        
         
    with open(file) as f:
        for i, j in enumerate(f):
            pass
     
    num = i + 1
         
    with open(file) as f:        
        for line in itertools.islice(f, num-line1, num-line2):
            if p == pattern_txt: 
                match = p.search(line)
                if match:
                    data = match.groupdict()['flt']
                    raw_data = '{0:.2f}'.format(float(data))
                    lists(folder_name, tool, data_src, _spec=spec, _field='-', _measure=raw_data)
     
            elif p == pattern_score:
                match = p.search(line)
                if match:
                    score_data = match.groupdict()['score']
                    lists(folder_name, tool, data_src, _spec=spec, _field='-', _measure=score_data)
 
 
def parseTxtPor(folder_name, tool, file, pattern, spec):
     
    data_src = data_source(file)
     
    with open(file) as f:
        for line in f:
            match = pattern.search(line)
            if match:
                data = match.groupdict()['por']
                value = int(data)/1000
                lists(folder_name, tool, data_src, _spec=spec, _field='tBOOTUP', _measure=value)
                 
             
def parseXlsx(folder_name, tool, file, pattern, race=file):
     
    data_src = data_source(file)
     
    wb = load_workbook(filename=file)
       
    if 'CDM301' in globals() and race == CDM301:
        sheet_ranges = wb['Sheet1']
    elif 'CDM503' in globals() and  race == CDM503:
        sheet_ranges = wb['Sheet1']
    elif 'CDM_1' in globals() and  race == CDM_1:
        sheet_ranges = wb['Sheet1']
    elif 'CDM_2' in globals() and  race == CDM_2:
        sheet_ranges = wb['Sheet1']
    elif 'CDM_3' in globals() and  race == CDM_3:
        sheet_ranges = wb['Sheet1']
    elif 'PCMARK_VANTAGE' in globals() and   race == PCMARK_VANTAGE:
        sheet_ranges = wb['Result 1']
     
    pcmark_lists = [140, 142, 144, 146, 148, 150, 152, 154, 139]
     
    if 'CDM301' in globals() and race == CDM301:
        iCDM301 = 0
        for row in sheet_ranges.iter_rows():
            for cell in row:
                if cell.value is not None:
                    match =  pattern.search(cell.value)    
                    if match:
                        iCDM301 = iCDM301 + 1
                        if iCDM301 == 1:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Read', _measure=match.groupdict()['num'])
                        elif iCDM301 == 2:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Write', _measure=match.groupdict()['num'])
                        elif iCDM301 == 3:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 512KB', _measure=match.groupdict()['num'])
                        elif iCDM301 == 4:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 512KB', _measure=match.groupdict()['num'])
                        elif iCDM301 == 5:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 4KB (QD=1)', _measure=match.groupdict()['num'])
                        elif iCDM301 == 6:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 4KB (QD=1)', _measure=match.groupdict()['num'])
                        elif iCDM301 == 7:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 4KB (QD=32)', _measure=match.groupdict()['num'])
                        elif iCDM301 == 8:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 4KB (QD=32)', _measure=match.groupdict()['num'])                    
               
    elif 'CDM503' in globals() and race == CDM503:
        iCDM503 = 0
        for row in sheet_ranges.iter_rows():  
            for cell in row:
                if cell.value is not None:
                    match =  pattern.search(cell.value)      
                    if match:
                        iCDM503 = iCDM503 + 1
                        if iCDM503 == 1:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Read (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 2:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Write (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 3:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 4:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 5:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='REad (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 6:
                            lists(folder_name, tool, data_src, _spec='Sequential', _field='Write (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM503 == 7:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Read 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num'])   
                        elif iCDM503 == 8:
                            lists(folder_name, tool, data_src, _spec='Random', _field='Write 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num']) 
    elif 'CDM_1' in globals() and race == CDM_1:
        iCDM_1 = 0
        for row in sheet_ranges.iter_rows():  
            for cell in row:
                if cell.value is not None:
                    match =  pattern.search(cell.value)      
                    if match:
                        iCDM_1 = iCDM_1 + 1
                        if iCDM_1 == 1:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src ,_field='Read (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_1 == 2:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src ,_field='Write (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_1 == 3:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Read 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_1 == 4:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Write 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_1 == 5:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='REad (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_1 == 6:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='Write (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_1 == 7:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Read 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num'])   
                        elif iCDM_1 == 8:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Write 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num']) 
    elif 'CDM_2' in globals() and race == CDM_2:
        iCDM_2 = 0
        for row in sheet_ranges.iter_rows():  
            for cell in row:
                if cell.value is not None:
                    match =  pattern.search(cell.value)      
                    if match:
                        iCDM_2 = iCDM_2 + 1
                        if iCDM_2 == 1:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='Read (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_2 == 2:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='Write (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_2 == 3:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Read 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_2 == 4:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Write 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_2 == 5:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='REad (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_2 == 6:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='Write (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_2 == 7:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Read 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num'])   
                        elif iCDM_2 == 8:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Write 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num'])  
    elif 'CDM_3' in globals() and race == CDM_3:
        iCDM_3 = 0
        for row in sheet_ranges.iter_rows():  
            for cell in row:
                if cell.value is not None:
                    match =  pattern.search(cell.value)      
                    if match:
                        iCDM_3 = iCDM_3 + 1
                        if iCDM_3 == 1:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='Read (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_3 == 2:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='Write (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_3 == 3:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Read 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_3 == 4:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Write 4KiB (Q= 32,T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_3 == 5:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='REad (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_3 == 6:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Sequential', _spec1=data_src , _field='Write (T= 1)', _measure=match.groupdict()['num'])
                        elif iCDM_3 == 7:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Read 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num'])   
                        elif iCDM_3 == 8:
                            lists1(folder_name, tool, 'CDM_PCI', _spec='Random', _spec1=data_src , _field='Write 4KiB (Q= 1,T= 1)', _measure=match.groupdict()['num'])                   
    elif 'PCMARK_VANTAGE' in globals() and  race == PCMARK_VANTAGE:
        for i in pcmark_lists:
            if i == 140:
                lists(folder_name, tool, data_src, _spec='HDD1', _field='Windows Defender', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 142:
                lists(folder_name, tool, data_src, _spec='HDD2', _field='Gaming', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 144:
                lists(folder_name, tool, data_src, _spec='HDD3', _field='Importing pictures to Windows Photo Gallery', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 146:
                lists(folder_name, tool, data_src, _spec='HDD4', _field='Windows Vista startup', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 148:
                lists(folder_name, tool, data_src, _spec='HDD5', _field='Video editing using Windows Movie Maker', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 150:
                lists(folder_name, tool, data_src, _spec='HDD6', _field='Windows Media Center', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 152:
                lists(folder_name, tool, data_src, _spec='HDD7', _field='Adding music to Windows Media Player', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 154:
                lists(folder_name, tool, data_src, _spec='HDD8', _field='Application loading', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 139:
                lists(folder_name, tool, data_src, _spec='Score', _field='-', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
    elif 'PCMARK_VANTAGE' in globals() and  race == PCMARK_VANTAGE:
        for i in pcmark_lists:
            if i == 140:
                lists(folder_name, tool, data_src, _spec='HDD1', _field='Windows Defender', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 142:
                lists(folder_name, tool, data_src, _spec='HDD2', _field='Gaming', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 144:
                lists(folder_name, tool, data_src, _spec='HDD3', _field='Importing pictures to Windows Photo Gallery', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 146:
                lists(folder_name, tool, data_src, _spec='HDD4', _field='Windows Vista startup', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 148:
                lists(folder_name, tool, data_src, _spec='HDD5', _field='Video editing using Windows Movie Maker', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 150:
                lists(folder_name, tool, data_src, _spec='HDD6', _field='Windows Media Center', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 152:
                lists(folder_name, tool, data_src, _spec='HDD7', _field='Adding music to Windows Media Player', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 154:
                lists(folder_name, tool, data_src, _spec='HDD8', _field='Application loading', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
            elif i == 139:
                lists(folder_name, tool, data_src, _spec='Score', _field='-', _measure=math.trunc(sheet_ranges['C' + str(i)].value))

def parseXlsxPCI_PcmarkVantage(folder_name, tool, file):
     
    data_src = data_source(file)
    wb = load_workbook(filename=file)
    sheet_ranges = wb['Result 1']
    pcmark_lists = [140, 142, 144, 146, 148, 150, 152, 154, 139]
     
    v_pattern = re.compile(r'PCMV(_(?P<gb>[\d]+))?')
    v_match = v_pattern.search(data_src)
    vGB = ""
    if v_match:
        if v_match.groupdict()['gb'] is not None:
            vGB = v_match.groupdict()['gb']
    for i in pcmark_lists:
        if i == 140:
            lists1(folder_name, tool, "PCMV_PCI", _spec='HDD1',_spec1=vGB, _field='Windows Defender', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
        elif i == 142:
            lists1(folder_name, tool, "PCMV_PCI", _spec='HDD2',_spec1=vGB, _field='Gaming', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
        elif i == 144:
            lists1(folder_name, tool, "PCMV_PCI", _spec='HDD3',_spec1=vGB, _field='Importing pictures to Windows Photo Gallery', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
        elif i == 146:
            lists1(folder_name, tool, "PCMV_PCI", _spec='HDD4',_spec1=vGB, _field='Windows Vista startup', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
        elif i == 148:
            lists1(folder_name, tool, "PCMV_PCI", _spec='HDD5',_spec1=vGB, _field='Video editing using Windows Movie Maker', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
        elif i == 150:
            lists1(folder_name, tool, "PCMV_PCI", _spec='HDD6',_spec1=vGB, _field='Windows Media Center', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
        elif i == 152:
            lists1(folder_name, tool, "PCMV_PCI", _spec='HDD7',_spec1=vGB, _field='Adding music to Windows Media Player', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
        elif i == 154:
            lists1(folder_name, tool, "PCMV_PCI", _spec='HDD8',_spec1=vGB, _field='Application loading', _measure=math.trunc(sheet_ranges['C' + str(i)].value))
        elif i == 139:
            lists1(folder_name, tool, "PCMV_PCI", _spec='Score',_spec1=vGB, _field='-', _measure=math.trunc(sheet_ranges['C' + str(i)].value))



def parsSpecWpcHtml(vfile):
    #resp = urllib.urlretrieve('file:///' + vfile['filePath'], 'resultHTML.html')
    #emp = resp
    v_web = urllib.urlopen('file:///' + vfile['filePath'])
    v_s = v_web.read()
    
    v_html = etree.HTML(v_s)
    
    ## Get all 'tr'
    #tr_nodes = html.xpath('//table[@id="Report1_dgReportDemographic"]/tr')
    v_table_nodes = v_html.xpath('//div[@class="tabContent"]/table')
    for v_table in v_table_nodes:
        v_title = v_table.getparent().xpath("h2")[0].text
        if v_title == 'Summary' :
            parsSpecSummary(vfile,v_table)
            
## only summary 
#         else :
#             v_tr_nodes = v_table.xpath("tr")
#             v_header = [] # [i[0].html() for i in v_tr_nodes[0].xpath("td")]
#             for i in v_tr_nodes[0].xpath("td"):
#                 v_header.append(i[0].text)
#             
#             v_contents = [] # [[td.html() for td in tr.xpath('td')] for tr in v_tr_nodes[1:]]
#             for tr in v_tr_nodes[1:] :
#                 v_content = []
#                 v_contents.append(v_content)
#                 for td in tr.xpath('td') :
#                     v_contents.append(td.text)
    
    
def parsSpecSummary(vfile,v_table):
    v_title = v_table.getparent().xpath("h2")[0].text
        
    v_tr_nodes = v_table.xpath("tr")
#     v_header = [] # [i[0].html() for i in v_tr_nodes[0].xpath("td")]
#     for i in v_tr_nodes[0].xpath("td"):
#         v_header.append(i[0].text)
    v_dictionary = {}    
    v_contents = [] # [[td.html() for td in tr.xpath('td')] for tr in v_tr_nodes[1:]]
    
    # 1/2
    for tr in v_tr_nodes  :
        v_td_field = tr.xpath("td")[0]
        v_td_value = tr.xpath("td")[1]
        
        if v_td_field.text == None and v_td_field.getchildren().__len__() == 0:
            continue
        
        if v_td_field.text != None and v_td_field.text.strip() == '' :
            continue
        
        if v_td_field.getchildren().__len__() > 0 and v_td_field[0].tag == 'h3' :
            v_dictionary = {}
            matrix.append(v_dictionary)
            
            v_dictionary["meta"] = {}
            v_dictionary["meta"]["fileName"] = 'SPECwpc'
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'SPECwpc'
            
            v_dictionary["keys"] = {}
            v_dictionary["keys"]["Access Specification Name"] = v_td_field[0].text
            v_dictionary["keys"]["Queue Depth"] = 'Summary' # ggomsu
            
            
            v_dictionary[v_td_field[0].text] = v_td_value[0].text
            
    
        else :
            v_dictionary[v_td_field.text] = v_td_value.text
            
    # 2/2
    for tr in v_tr_nodes  :
        v_td_field = tr.xpath("td")[3]
        v_td_value = tr.xpath("td")[4]
        
        # followed code is same with 1/2
        if v_td_field.text == None and v_td_field.getchildren().__len__() == 0:
            continue
        
        if v_td_field.text != None and v_td_field.text.strip() == '' :
            continue
        
        if v_td_field.getchildren().__len__() > 0 and v_td_field[0].tag == 'h3' :
            v_dictionary = {}
            matrix.append(v_dictionary)
            
            v_dictionary["meta"] = {}
            v_dictionary["meta"]["fileName"] = 'SPECwpcSummary'
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'SPECwpc'
            
            v_dictionary["keys"] = {}
            v_dictionary["keys"]["Access Specification Name"] = v_td_field[0].text
            v_dictionary["keys"]["Queue Depth"] = 'Summary' # ggomsu
            
            
            v_dictionary[v_td_field[0].text] = v_td_value[0].text
            
    
        else :
            v_dictionary[v_td_field.text] = v_td_value.text
        
def parseWpcCsv(vfile):
    v_f = open(vfile['filePath'], 'r')
    #folderInfo
    v_csvReader = csv.reader(v_f)
    v_header = []
     
    for row in v_csvReader:
        if row[0] == "'Target Type" and len(v_header) == 0 :
            v_header = row
        
        if row[0] ==  "MANAGER" :
            v_dictionary = dict(zip(v_header, row))
            
          
            v_dictionary["meta"] = {}
            v_dictionary["meta"]["fileName"] = vfile['fileName'] 
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'SPECwpc'
             
            v_dictionary["keys"] = {}
            v_dictionary["keys"]["Access Specification Name"] = v_dictionary["Access Specification Name"]
            del v_dictionary["Access Specification Name"]
            v_dictionary["keys"]["Queue Depth"] = v_dictionary["Queue Depth"]
            del v_dictionary["Queue Depth"]
             
            matrix.append(v_dictionary)
#             print dictionary
       
    v_f.close()  
    
def parseIometerRWBlockSweep(vfile):
    v_f = open(vfile['filePath'], 'r')
    #folderInfo
    v_csvReader = csv.reader(v_f)
    v_header = []
     
    for row in v_csvReader:
        if len(row) == 0:
            continue
        if row[0] == "Block Size":
            v_header = row
            i = 0
            for h in v_header:
                v_header[i] = h.strip()
                i = i + 1
            continue
        if len(v_header) == 0 :
            continue
        
        v_dictionary = dict(zip(v_header, row))
        dicValStrip(v_dictionary)
      
        v_dictionary["meta"] = {}
        v_dictionary["meta"]["fileName"] = vfile['fileName'] #"Max Throughput.csv"
        v_dictionary["meta"]["folderName"] = vfile['folderName']
        v_dictionary["meta"]["tool"] = 'Iometer'
         
        v_dictionary["keys"] = {}
        v_dictionary["keys"]["Access Specification Name"] = v_dictionary["Random vs. Sequential"].strip() + "_" + v_dictionary["Read vs. Write"].strip() + "_" + v_dictionary["Block Size"].strip()
        #del dictionary["Access Specification Name"]
        v_dictionary["keys"]["Queue Depth"] = "undefined"
        #del dictionary["Queue Depth"]
         
        matrix.append(v_dictionary)
#             print dictionary
       
    v_f.close()
    
def parseIometerPciQd(vfile):
    vfileSplit = vfile['fileName'].split(" ")
    #vSpec = vfileSplit[1] + "_" + vfileSplit[2] + "_" + vfileSplit[3]
    
    pattern_inst = re.compile(r'QD(?P<qd>[\d]+)\s[\d]+KB(_(?P<gb>[\d]+GB))?')
    match = pattern_inst.search(vfile['fileName'])
    vGB = ""
    vQD = ""
    if match:
        vQD = match.groupdict()['qd']
        if match.groupdict()['gb'] is not None:
            vGB = match.groupdict()['gb']
    
    
    v_f = open(vfile['filePath'], 'r')
    
    #folderInfo
    v_csvReader = csv.reader(v_f)
    v_header = []
     
    for row in v_csvReader:
        if len(row) > 1 and row[1] == "'Target Type" and len(v_header) == 0 :
            v_header = row
        
        if len(row) > 1 and row[1] ==  "ALL" :        
            v_dictionary = dict(zip(v_header, row))
            dicValStrip(v_dictionary)
          
            v_dictionary["meta"] = {}
            # v_dictionary["meta"]["fileName"] = file['fileName'] #"Max Throughput.csv"
            v_dictionary["meta"]["fileName"] = 'QD_PCI' + vGB
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'Iometer'
             
            v_dictionary["keys"] = {}
            v_dictionary["keys"]["Access Specification Name"] = v_dictionary["Access Specification Name"]
            #v_dictionary["keys"]["spec1"] = v_dictionary['TimeStamp']
            del v_dictionary["Access Specification Name"]
            #v_dictionary["keys"]["Queue Depth"] = v_dictionary["Queue Depth"]
            v_dictionary["keys"]["Queue Depth"] = vQD
            del v_dictionary["Queue Depth"]
             
            matrix.append(v_dictionary)
#             print dictionary
       
    v_f.close()    
def parseWorderPci(vfile):
    vfileSplit = vfile['fileName'].split(" ")
    #vSpec = vfileSplit[1] + "_" + vfileSplit[2] + "_" + vfileSplit[3]
    
    pattern_inst = re.compile(r'(?P<worker>[\d]+)W_QD(?P<qd>[\d]+)(_(?P<gb>[\d]+GB))?')
    match = pattern_inst.search(vfile['fileName'])
    vWorder = ""
    vGB = ""
    vQD = ""
    if match:
        vWorder = match.groupdict()['worker']
        vQD = match.groupdict()['qd']
        if match.groupdict()['gb'] is not None:
            vGB = match.groupdict()['gb']
    
    
    v_f = open(vfile['filePath'], 'r')
    
    #folderInfo
    v_csvReader = csv.reader(v_f)
    v_header = []
     
    for row in v_csvReader:
        if len(row) > 1 and row[1] == "'Target Type" and len(v_header) == 0 :
            v_header = row
        
        if len(row) > 1 and row[1] ==  "ALL" :        
            v_dictionary = dict(zip(v_header, row))
            dicValStrip(v_dictionary)
          
            v_dictionary["meta"] = {}
            # v_dictionary["meta"]["fileName"] = file['fileName'] #"Max Throughput.csv"
            v_dictionary["meta"]["fileName"] = 'WORKER_PCI' + vGB
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'Iometer'
             
            v_dictionary["keys"] = {}
            v_dictionary["keys"]["Access Specification Name"] = v_dictionary["Access Specification Name"]
            v_dictionary["keys"]["spec1"] = vWorder
            del v_dictionary["Access Specification Name"]
            #v_dictionary["keys"]["Queue Depth"] = v_dictionary["Queue Depth"]
            v_dictionary["keys"]["Queue Depth"] = vQD
            del v_dictionary["Queue Depth"]
             
            matrix.append(v_dictionary)
            print v_dictionary
       
    v_f.close()   

def parsePciLatency(vfile):
    vfileSplit = vfile['fileName'].split(" ")
    #vSpec = vfileSplit[1] + "_" + vfileSplit[2] + "_" + vfileSplit[3]
#     pattern_inst = re.compile(r'QD([\d]+)\s[\d]+KB_(\dGB)')
#     
#     vMatch = pattern_inst.findall(vfile['fileName'])
#     vGB = ""
#     if len(vMatch) > 0 :
#         vGB = vMatch[1]
    pattern_inst = re.compile(r'QD(?P<qd>[\d]+)\s[\d]+KB(_(?P<gb>[\d]+GB))?')
    match = pattern_inst.search(vfile['fileName'])
    vGB = ""
    vQD = ""
    if match:
        vQD = match.groupdict()['qd']
        if match.groupdict()['gb'] is not None:
            vGB = match.groupdict()['gb']
    
    v_f = open(vfile['filePath'], 'r')
    
    #folderInfo
    v_csvReader = csv.reader(v_f)
    v_header = []
     
    for row in v_csvReader:
        if len(row) > 1 and row[1] == "'Target Type" and len(v_header) == 0 :
            v_header = row
        
        if len(row) > 1 and row[1] ==  "ALL" :        
            v_dictionary = dict(zip(v_header, row))
            dicValStrip(v_dictionary)
          
            v_dictionary["meta"] = {}
            # v_dictionary["meta"]["fileName"] = file['fileName'] #"Max Throughput.csv"
            v_dictionary["meta"]["fileName"] = 'Latency_PCI' + vGB
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'Iometer'
             
            v_dictionary["keys"] = {}
            v_dictionary["keys"]["Access Specification Name"] = v_dictionary["Access Specification Name"]
            #v_dictionary["keys"]["spec1"] = v_dictionary['TimeStamp']
            del v_dictionary["Access Specification Name"]
            #v_dictionary["keys"]["Queue Depth"] = v_dictionary["Queue Depth"]
            v_dictionary["keys"]["Queue Depth"] = vQD
            del v_dictionary["Queue Depth"]
             
            matrix.append(v_dictionary)
#             print dictionary
       
    v_f.close()    
 
 
def dicValStrip(vDic):
    for d in  vDic:
        if type(vDic[d]) is str:
            vDic[d] = vDic[d].strip() 
        
def dicValRemoveSpecialString(vDic):
    for d in  vDic:
        if type(vDic[d]) is str and "'" in d :
            vCol = d.replace("'","")
            vDic[vCol] = vDic[d].strip()
            del vDic[d]
          
def parseSustain(vfile):
    vfileSplit = vfile['fileName'].split(" ")
    vSpec = vfileSplit[1] + "_" + vfileSplit[2] + "_" + vfileSplit[3]
    v_f = open(vfile['filePath'], 'r')
    #folderInfo
    v_csvReader = csv.reader(v_f)
    v_header = []
     
    for row in v_csvReader:
        if len(row) > 1 and row[1] == "'Target Type" and len(v_header) == 0 :
            v_header = row
        
        if len(row) > 1 and row[1] ==  "WORKER" :        
            v_dictionary = dict(zip(v_header, row))
            dicValStrip(v_dictionary)
          
            v_dictionary["meta"] = {}
            # v_dictionary["meta"]["fileName"] = file['fileName'] #"Max Throughput.csv"
            v_dictionary["meta"]["fileName"] = 'Sustain'
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'Sustain'
             
            v_dictionary["keys"] = {}
            v_dictionary["keys"]["Access Specification Name"] = vSpec
            v_dictionary["keys"]["spec1"] = v_dictionary['TimeStamp']
            del v_dictionary["Access Specification Name"]
            v_dictionary["keys"]["Queue Depth"] = v_dictionary["Queue Depth"]
            del v_dictionary["Queue Depth"]
             
            matrix.append(v_dictionary)
#             print dictionary
       
    v_f.close()        
   
   
# PCI Sustain     
def parseSustained(vfile):
    vfileSplit = vfile['fileName'].split(" ")
    vSpec = vfileSplit[1] + "_" + vfileSplit[2] + "_" + vfileSplit[3]
    v_f = open(vfile['filePath'], 'r')
    #folderInfo
    v_csvReader = csv.reader(v_f)
    v_header = []
     
    for row in v_csvReader:
        if len(row) > 1 and row[1] == "'Target Type" and len(v_header) == 0 :
            v_header = row
        
        if len(row) > 1 and row[1] ==  "WORKER" :        
            v_dictionary = dict(zip(v_header, row))
            dicValStrip(v_dictionary)
          
            v_dictionary["meta"] = {}
            # v_dictionary["meta"]["fileName"] = file['fileName'] #"Max Throughput.csv"
            v_dictionary["meta"]["fileName"] = 'Sustained'
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'Sustain'
             
            v_dictionary["keys"] = {}
            v_dictionary["keys"]["Access Specification Name"] = vSpec
            v_dictionary["keys"]["spec1"] = v_dictionary['TimeStamp']
            del v_dictionary["Access Specification Name"]
            v_dictionary["keys"]["Queue Depth"] = v_dictionary["Queue Depth"]
            del v_dictionary["Queue Depth"]
             
            matrix.append(v_dictionary)
#             print dictionary
       
    v_f.close()        
    
    
def parseSSDPerformanceV1(vfile):
    v_f = open(vfile['filePath'], 'r')
    #folderInfo
    v_csvReader = csv.reader(v_f)
    v_header = []
     
    vSeq = 0
    vNum = []
    for row in v_csvReader:
        if row[0] == "'Target Type" and len(v_header) == 0 :
            v_header = row
        
        if row[0] ==  "MANAGER" :
        
            v_dictionary = dict(zip(v_header, row))
            dicValStrip(v_dictionary)
          
            v_dictionary["meta"] = {}
            v_dictionary["meta"]["fileName"] = vfile['fileName'] #"Max Throughput.csv"
            v_dictionary["meta"]["folderName"] = vfile['folderName']
            v_dictionary["meta"]["tool"] = 'Iometer'
             
            v_dictionary["keys"] = {}
            if v_dictionary["Access Specification Name"] not in vNum :
                vNum.append(v_dictionary["Access Specification Name"])
            else :
                vNum = [v_dictionary["Access Specification Name"]]
                vSeq = vSeq + 1
            
            v_dictionary["keys"]["spec1"] = str(vSeq)
            v_dictionary["keys"]["Access Specification Name"] = v_dictionary["Access Specification Name"]
            del v_dictionary["Access Specification Name"]
            v_dictionary["keys"]["Queue Depth"] = v_dictionary["Queue Depth"]
            del v_dictionary["Queue Depth"]
             
            matrix.append(v_dictionary)
#             print dictionary
       
    v_f.close()        
    

        
class MyError(Exception):
    pass    

#Iometer start
subprocess.call(r'net use z: /del', shell=True)
subprocess.call(r'net use z: \\10.15.202.120\das_input /user:hynixad\probe test')

matrix = []
vToday = datetime.datetime.now().strftime("%Y%m%d")
query = ''
#path = "//10.15.202.120/das_input/Solution/SSD_Performance"
# path = "//10.15.202.120/das_input/DAS_SOL/SSD_Performance"
path = "D:/hynix/requirement/Performance/PCI/test"
dirsIometer = os.listdir(path);
# folderHeaders = ['SYSTEM', 'OPERATING_SYSTEM', 'VENDOR', 'PRODUCT_NAME', 'NAND_TECH', 'CELL_TYPE','CONTROLLER', 'FORM_FACTOR', 'CAPACITY', 'FIRMWARE', 'SLC_BUFFER', 'MEASURE_DT']
folderHeaders = ['MEASURE_DT','VENDOR','PRODUCT_NAME', 'CONTROLLER', 'NAND_TECH','CELL_TYPE', 'FORM_FACTOR','CAPACITY','FIRMWARE','SLC_BUFFER','SERIAL_NUMBER','TEST_COUNT']
gFolderInfos = []
filesIometer = []

for subdir in dirsIometer:
    print subdir
    if subdir == 'archive':
        continue
    
    #folderInfos
    gFolderInfos.append({"folderName":subdir,"data":dict(zip(folderHeaders, subdir.split("_")))});
        
    pathIometer = path + '/' + subdir + '/Iometer'
    print pathIometer
    if os.path.exists(pathIometer):
        dirIometer = os.listdir(pathIometer)
        
        for file in dirIometer:
            fileObj = {"filePath":pathIometer + '/' + file,"folderName":subdir,"fileName":file, "fileFolderName": 'Iometer'}
            filesIometer.append(fileObj)
    
	#Latency 2017.09.20
	pathLatency = path + '/' + subdir + '/Latency'
    print pathLatency
    if os.path.exists(pathLatency):
        dirIometer = os.listdir(pathLatency)
        for file in dirIometer:
            fileObj = {"filePath":pathIometer + '/' + file,"folderName":subdir,"fileName":file, "fileFolderName": 'Latency'}
            filesIometer.append(fileObj)
			
    # SPECwpc 2017.08.24
    pathSPECwpc = path + '/' + subdir + '/SPECwpc'
    print pathSPECwpc
    if os.path.exists(pathSPECwpc):
        dirSPECwpc = os.listdir(pathSPECwpc)
        
        for file in dirSPECwpc:
            fileObj = {"filePath":pathSPECwpc + '/' + file,"folderName":subdir,"fileName":file, "fileFolderName": 'SPECwpc'}
            filesIometer.append(fileObj)
    # SPECwpc End
    
    # Sustain 2017.09.05
    pathSustain = path + '/' + subdir + '/Sustain'
    print pathSustain
    if os.path.exists(pathSustain):
        dirSustain = os.listdir(pathSustain)
        
        for file in dirSustain:
            fileObj = {"filePath":pathSustain + '/' + file,"folderName": subdir,"fileName":file, "fileFolderName": 'Sustain'}
            filesIometer.append(fileObj)
            
    # Sustain 2017.09.05 pcie
    pathSustain = path + '/' + subdir + '/Sustained'
    print pathSustain
    if os.path.exists(pathSustain):
        dirSustain = os.listdir(pathSustain)
        
        for file in dirSustain:
            fileObj = {"filePath":pathSustain + '/' + file,"folderName": subdir,"fileName":file, "fileFolderName": 'Sustained'}
            filesIometer.append(fileObj)
    # Sustain End
    
    # Worker 2017.09.20 pcie
    pathWorker = path + '/' + subdir + '/Worker'
    print pathWorker
    if os.path.exists(pathWorker):
        dirWorker = os.listdir(pathWorker)
        
        for file in dirWorker:
            fileObj = {"filePath":pathWorker + '/' + file,"folderName": subdir,"fileName":file, "fileFolderName": 'Worker'}
            filesIometer.append(fileObj)

# 01.Max Throughput.csv
for file in filesIometer:
    extension = os.path.splitext(file['filePath'])
    # in iometer not manage xlsx 
    if extension[1] == '.xlsx':
        continue
    
    if extension[1] == '.html':
        parsSpecWpcHtml(file)
        continue
    
    if file['fileName'] == 'RW Block Sweep parsed.csv':
        parseIometerRWBlockSweep(file)
        continue
    
    if file['fileName'] == 'SSD Performance V1.csv':
        parseSSDPerformanceV1(file)
        continue
        
    if file['fileName'] == 'wpc1.csv' or file['fileName'] == 'wpc2.csv' or file['fileName'] == 'wpc3.csv':
        parseWpcCsv(file)       
        continue
    
    if file['fileFolderName'] == 'Sustain' :
        rx = re.compile("^inst")
        matches = rx.findall(file['fileName'])
        if len(matches) > 0 :
            parseSustain(file)
        continue
    
    if file['fileFolderName'] == 'Sustained' :
        rx = re.compile("^inst")
        matches = rx.findall(file['fileName'])
        if len(matches) > 0 :
            parseSustained(file)
        continue
    
    # iometer  inst  parsing
    if file['fileFolderName'] == 'Iometer' :
        rx = re.compile("^inst")
        matches = rx.findall(file['fileName'])
        if len(matches) > 0 :
            parseIometerPciQd(file)
            continue
    # iometer  inst  parsing
    if file['fileFolderName'] == 'Latency' :
        rx = re.compile("^inst")
        matches = rx.findall(file['fileName'])
        if len(matches) > 0 :
            parsePciLatency(file)
            continue
        
    if file['fileFolderName'] == 'Worker' :
        rx = re.compile("^inst")
        matches = rx.findall(file['fileName'])
        if len(matches) > 0 :
            parseWorderPci(file)
            continue
			
    f = open(file['filePath'], 'r')
     
#      d_reader = csv.DictReader(f)
#      headers = d_reader.fieldnames
#      for line in d_reader:
#          matrix.append(line)
     
    #folderInfo
    folderInfos = file['folderName'].split("_")
    dicFolder = dict(zip(folderHeaders, folderInfos))
    csvReader = csv.reader(f)
    header = []
     
    for row in csvReader:
        if row[0] == "'Target Type" and len(header) == 0 :
            header = row
        if row[0] ==  "WORKER" :
            dictionary = dict(zip(header, row))
            if dictionary["Access Specification Name"] == 'Precondition':
                continue
            # delete ' charater
            dictionary["Target Type"] = dictionary["'Target Type"]
            del dictionary["'Target Type"]
            dictionary["meta"] = {}
            dictionary["meta"]["fileName"] = file['fileName'] #"Max Throughput.csv"
            dictionary["meta"]["folderName"] = file['folderName']
            dictionary["meta"]["tool"] = 'Iometer'
             
            #header infos
            for head in folderHeaders:
                dictionary["meta"][head] = dicFolder[head]
             
            dictionary["keys"] = {}
            dictionary["keys"]["Access Specification Name"] = dictionary["Access Specification Name"]
            del dictionary["Access Specification Name"]
            dictionary["keys"]["Queue Depth"] = dictionary["Queue Depth"]
            del dictionary["Queue Depth"]
             
            matrix.append(dictionary)
#             print dictionary
       
    f.close()
# raise MyError
# print len(matrix)
try:
     
    vProduct = '' 
    #con_str = 'swdashboard/swdashboard@166.125.19.98:1521/APS'
    con_str = 'swdashboard/swdashboard@166.125.19.99:1521/APS'
    con = cx_Oracle.connect(con_str)
    cur = con.cursor()
      
    for folders in gFolderInfos:   
        i = 0
          
        query = "delete from PERFORMANCE_FOLDER where " 
        query += " FOLDER_NAME ='" + folders['folderName'] +"'"
        cur.execute(query)
          
        query = "delete from PERFORMANCE where " 
        query += " FOLDER_NAME ='" + folders['folderName'] +"'"
#         query += " and tool = 'Iometer'"
        cur.execute(query)
          
        query = "insert into PERFORMANCE_FOLDER(FOLDER_NAME, MEASURE_DT, VENDOR, PRODUCT_NAME, CONTROLLER, NAND_TECH, CELL_TYPE, FORM_FACTOR, CAPACITY, FIRMWARE, SLC_BUFFER, SERIAL_NUMBER, TEST_COUNT,CATEGORY)"
        query += " values("
        query += "'" + folders['folderName'] + "'"
        for head in folderHeaders:
            query += ",'" + folders['data'][head] + "'" 
            i = i + 1
        query += ",'CSSD_PCI'"       
        query += ")"
        cur.execute(query)
#     con.commit()   
          
          
    i = 0
    #rx = re.compile("^inst")
    for m in matrix:
        i = i+1
        print repr(i) + '(' + repr(len(matrix)) + ')'
        dicValStrip(m)
        dicValRemoveSpecialString(m)
        
        for mKey in m: 
            if type(m[mKey]) is dict:
                continue
            vParam = {}
            vParam['FOLDERNAME'] = str(m['meta']['folderName'])
            vParam['TOOL'] = str(m['meta']['tool'])
            vParam['FILENAME'] = str(m['meta']['fileName'])
            vParam['SPEC'] = str(m['keys']['Access Specification Name'])            
            vParam['QD'] = str(m['keys']['Queue Depth'])
            vParam['FIELD'] = str(mKey)
            vParam['MEASURE'] = str(m[mKey])
            query = ''
            
            if vParam['FILENAME'] in ('QD_PCI','Latency_PCI') :
                if m['keys'].get('spec1') is not None:
                    vParam['SPEC1'] = str(m['keys']['spec1'])            
                    query = """
                         DECLARE                
                            VCNT NUMBER;    
                        BEGIN
                            SELECT COUNT(*) INTO VCNT FROM PERFORMANCE
                            WHERE FOLDER_NAME = '""" + vParam['FOLDERNAME'] +"""'
                            AND TOOL = '""" + vParam['TOOL'] +"""'
                            AND DATA_SRC = '""" + vParam['FILENAME'] +"""' 
                            AND SPEC = '""" + vParam['SPEC'] +"""'
                            AND SPEC1 = '""" + vParam['SPEC1'] +"""'
                            AND QUEUE_DEPTH = '""" + vParam['QD'] +"""' 
                            AND FIELD = '""" + vParam['FIELD'] +"""' ; 
                             
                            IF VCNT = 0 THEN 
                                INSERT INTO PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC, SPEC1,QUEUE_DEPTH, FIELD, MEASURE)
                                VALUES ( 
                                            '""" + vParam['FOLDERNAME'] +"""' , 
                                            '""" + vParam['TOOL'] +"""' , 
                                            '""" + vParam['FILENAME'] +"""' , 
                                            '""" + vParam['SPEC'] +"""' , 
                                            '""" + vParam['SPEC1'] +"""' , 
                                            '""" + vParam['QD'] +"""' , 
                                            '""" + vParam['FIELD'] +"""' , 
                                            '""" + vParam['MEASURE'] +"""' 
                                        );
                     
                            ELSE 
                                update PERFORMANCE set
                                    measure = '""" + vParam['MEASURE'] +"""' 
                                WHERE FOLDER_NAME = '""" + vParam['FOLDERNAME'] +"""'
                                AND TOOL = '""" + vParam['TOOL'] +"""'
                                AND DATA_SRC = '""" + vParam['FILENAME'] +"""' 
                                AND SPEC = '""" + vParam['SPEC'] +"""'
                                AND SPEC1 = '""" + vParam['SPEC1'] +"""'
                                AND QUEUE_DEPTH = '""" + vParam['QD'] +"""' 
                                AND FIELD = '""" + vParam['FIELD'] +"""' ; 
                     
                            END IF;          
                    
                        END;
                    """    
                else:
                    query = """
                         DECLARE                
                            VCNT NUMBER;    
                        BEGIN
                            SELECT COUNT(*) INTO VCNT FROM PERFORMANCE
                            WHERE FOLDER_NAME = '""" + vParam['FOLDERNAME'] +"""'
                            AND TOOL = '""" + vParam['TOOL'] +"""'
                            AND DATA_SRC = '""" + vParam['FILENAME'] +"""' 
                            AND SPEC = '""" + vParam['SPEC'] +"""'
                            AND QUEUE_DEPTH = '""" + vParam['QD'] +"""' 
                            AND FIELD = '""" + vParam['FIELD'] +"""' ; 
                             
                            IF VCNT = 0 THEN 
                                INSERT INTO PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC, QUEUE_DEPTH, FIELD, MEASURE)
                                VALUES ( 
                                            '""" + vParam['FOLDERNAME'] +"""' , 
                                            '""" + vParam['TOOL'] +"""' , 
                                            '""" + vParam['FILENAME'] +"""' , 
                                            '""" + vParam['SPEC'] +"""' , 
                                            '""" + vParam['QD'] +"""' , 
                                            '""" + vParam['FIELD'] +"""' , 
                                            '""" + vParam['MEASURE'] +"""' 
                                        );
                     
                            ELSE 
                                update PERFORMANCE set
                                    measure = '""" + vParam['MEASURE'] +"""' 
                                WHERE FOLDER_NAME = '""" + vParam['FOLDERNAME'] +"""'
                                AND TOOL = '""" + vParam['TOOL'] +"""'
                                AND DATA_SRC = '""" + vParam['FILENAME'] +"""' 
                                AND SPEC = '""" + vParam['SPEC'] +"""'
                                AND QUEUE_DEPTH = '""" + vParam['QD'] +"""' 
                                AND FIELD = '""" + vParam['FIELD'] +"""' ; 
                     
                            END IF;          
                    
                        END;
                    """
            else:
                if m['keys'].get('spec1') is not None:
                    vParam['SPEC1'] = str(m['keys']['spec1'])            
                    query = """
                                INSERT INTO PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC, SPEC1,QUEUE_DEPTH, FIELD, MEASURE)
                                VALUES ( 
                                            '""" + vParam['FOLDERNAME'] +"""' , 
                                            '""" + vParam['TOOL'] +"""' , 
                                            '""" + vParam['FILENAME'] +"""' , 
                                            '""" + vParam['SPEC'] +"""' , 
                                            '""" + vParam['SPEC1'] +"""' , 
                                            '""" + vParam['QD'] +"""' , 
                                            '""" + vParam['FIELD'] +"""' , 
                                            '""" + vParam['MEASURE'] +"""' 
                                        )
                    """    
                else:
                    query = """
                                INSERT INTO PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC, QUEUE_DEPTH, FIELD, MEASURE)
                                VALUES ( 
                                            '""" + vParam['FOLDERNAME'] +"""' , 
                                            '""" + vParam['TOOL'] +"""' , 
                                            '""" + vParam['FILENAME'] +"""' , 
                                            '""" + vParam['SPEC'] +"""' , 
                                            '""" + vParam['QD'] +"""' , 
                                            '""" + vParam['FIELD'] +"""' , 
                                            '""" + vParam['MEASURE'] +"""' 
                                        )
                    
                    """
            
            cur.execute(query)
            #con.commit()
            #print query
            
            
#             #matches = rx.findall(m['fileName'])
#             if m['meta']['fileName'] in ('wpc1.csv','wpc2.csv','wpc3.csv'):
#             #if len(matches) > 0 :
#                 if mKey in ('MBps (Decimal)', 'Read MBps (Decimal)','Write MBps (Decimal)')  :
#                     query = " insert into PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC, SPEC_LABEL, QUEUE_DEPTH, FIELD, MEASURE)"
#                     query += " values('" + m["meta"]['folderName'] +"','" + m["meta"]['tool'] +"','" + m["meta"]['fileName'] +"', '" + m["keys"]['Access Specification Name'] +"','', '" + m["keys"]["Queue Depth"] +"', '"+ mKey +"',  '" + m[mKey] + "'  )"
#             elif m['meta']['fileName'] in ('SSD Performance V1.csv'):
#             #if len(matches) > 0 :
#                 if mKey in ( 'MBps (Decimal)','IOps') :
#                     query = " insert into PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC,SPEC1, SPEC_LABEL, QUEUE_DEPTH, FIELD, MEASURE)"
#                     query += " values('" + m["meta"]['folderName'] +"','" + m["meta"]['tool'] +"','" + m["meta"]['fileName'] +"', '" + m["keys"]['Access Specification Name'] +"', '" + m["keys"]['spec1'] +"'" + ",'', '" + m["keys"]["Queue Depth"] +"', '"+ mKey +"',  '" + m[mKey] + "'  )"
#             elif m['meta']['fileName'] in ('Sustain'):
#             #if len(matches) > 0 :
#                 if mKey in ('MBps (Decimal)') :
#                     query = " insert into PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC,SPEC1, SPEC_LABEL, QUEUE_DEPTH, FIELD, MEASURE)"
#                     query += " values('" + m["meta"]['folderName'] +"','" + m["meta"]['tool'] +"','" + m["meta"]['fileName'] +"', '" + m["keys"]['Access Specification Name'] +"', '" + m["keys"]['spec1'] +"'" + ",'', '" + m["keys"]["Queue Depth"] +"', '"+ mKey +"',  '" + m[mKey] + "'  )"
#             else :
#                 query = " insert into PERFORMANCE(FOLDER_NAME, TOOL, DATA_SRC, SPEC, SPEC_LABEL, QUEUE_DEPTH, FIELD, MEASURE)"
#                 query += " values('" + m["meta"]['folderName'] +"','" + m["meta"]['tool'] +"','" + m["meta"]['fileName'] +"', '" + m["keys"]['Access Specification Name'] +"','', '" + m["keys"]["Queue Depth"] +"', '"+ mKey +"',  '" + m[mKey] + "'  )"
#                
#             if query.__len__() > 0:
#                 cur.execute(query)
#                 print query
    con.commit()
    
    
    #Iometer End
    #not Iometer
    # Tag
    seq_test_r_tag = 'SeqTest/Read'
    seq_test_w_tag = 'SeqTest/Write'
    Random4K1TTest_r_tag = 'Random4K1TTest/Read'
    Random4K1TTest_w_tag = 'Random4K1TTest/Write'        
    Random4K64TTest_r_tag = 'Random4K64TTest/Read'
    Random4K64TTest_w_tag = 'Random4K64TTest/Write'     
    acc_time_r_tag = 'AccTimeTest/Read'
    acc_time_w_tag = 'AccTimeTest/Write'
    overall_tag = '{http://www.bapco.com}overall'
    scenario_tag = '{http://www.bapco.com}scenario' 
    overallperformance_tag = '{http://www.bapco.com}data/{http://www.bapco.com}overallperformance'
    storage_score = 'results/result/storage_score'
    storage_bandwidth = 'results/result/storage_bandwidth'
     
    # Regex
    pattern = re.compile(r'((?P<num>\d*).\d+)\sMB\/s')
    pattern_txt = re.compile(r'((?P<flt>\d*.\d+))\sMB\/s')
    pattern_score = re.compile(r'(score:\s(?P<score>\d+)\s)')
    pattern_tBOOTUP = re.compile(r'tBOOTUP\s:\s(?P<por>\d+)')
                 
    #pwd = '//10.15.202.120/das_input/DAS_SOL/SSD_Performance'
    #pwd = "D:/hynix/requirement/Performance/parsingFile_new/Test1"
    pwd = path
    for path, dirs, files in os.walk(pwd):
        dirs[:] = [dir for dir in dirs if dir not in ("Iometer","SPECwpc","Sustain","Sustained","Latency","Worker")  and dir not in("archive") ]
        for file in files:
            if os.path.basename(file) != 'readme.txt':
                current_dir_path = os.path.join(path, file)             
                path_sep = os.path.sep
                components = current_dir_path.split(path_sep)                    
                 
                rx = re.compile("^PCMARK_VANTAGE")
                matches = rx.findall(file)
                
                v_rx_pci_pcmv = re.compile("^PCMV")
                v_matches_pci_pcmv = v_rx_pci_pcmv.findall(file)
                if len(matches) > 0 :
                #if file == 'PCMARK_VANTAGE.xlsx':
                    PCMARK_VANTAGE = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, PCMARK_VANTAGE, pattern, race=PCMARK_VANTAGE)
                elif len(v_matches_pci_pcmv) > 0 :
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseXlsx'
                    parseXlsxPCI_PcmarkVantage(folder_name, tool, current_dir_path)
                     
                elif file == 'CDM301.xlsx':
                    CDM301 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]        
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, CDM301, pattern, race=CDM301)   
                     
                elif file == 'CDM503.xlsx':
                    CDM503 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, CDM503, pattern, race=CDM503)
                   
                elif file == 'CDM_1.xlsx':
                    CDM_1 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, CDM_1, pattern, race=CDM_1) 
                elif file == 'CDM_2.xlsx':
                    CDM_2 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, CDM_2, pattern, race=CDM_2) 
                elif file == 'CDM_3.xlsx':
                    CDM_3 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseXlsx'
                    parseXlsx(folder_name, tool, CDM_3, pattern, race=CDM_3) 
                elif file == 'AS-SSD1.6.xml':
                    AS_SSDv16 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]                        
                    print '----AS-SSDv1.6'
                    parseXml(folder_name, tool, AS_SSDv16, seq_test_r_tag)
                    parseXml(folder_name, tool, AS_SSDv16, seq_test_w_tag)
                    parseXml(folder_name, tool, AS_SSDv16, Random4K1TTest_r_tag)
                    parseXml(folder_name, tool, AS_SSDv16, Random4K1TTest_w_tag)
                    parseXml(folder_name, tool, AS_SSDv16, Random4K64TTest_r_tag)
                    parseXml(folder_name, tool, AS_SSDv16, Random4K64TTest_w_tag)
                    parseAccTime(folder_name, tool, AS_SSDv16, acc_time_r_tag)
                    parseAccTime(folder_name, tool, AS_SSDv16, acc_time_w_tag)
                     
                elif file == 'AS-SSD1.8.xml':
                    AS_SSDv18 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----AS-SSDv1.8'
                    parseXml(folder_name, tool, AS_SSDv18, seq_test_r_tag)
                    parseXml(folder_name, tool, AS_SSDv18, seq_test_w_tag)           
                    parseXml(folder_name, tool, AS_SSDv18, Random4K1TTest_r_tag)
                    parseXml(folder_name, tool, AS_SSDv18, Random4K1TTest_w_tag)
                    parseXml(folder_name, tool, AS_SSDv18, Random4K64TTest_r_tag)
                    parseXml(folder_name, tool, AS_SSDv18, Random4K64TTest_w_tag)
                    parseAccTime(folder_name, tool, AS_SSDv18, acc_time_r_tag)
                    parseAccTime(folder_name, tool, AS_SSDv18, acc_time_w_tag)
                     
                elif file == 'PCMARK7.txt':
                    PCMARK7 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseTxt'
                    parseTxt(folder_name, tool, PCMARK7, 21, 20, 'Windows Defender', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 20, 19, 'Importing pictures', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 19, 18, 'Video editing', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 18, 17, 'Windows Media Center', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 17, 16, 'Adding Music', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 16, 15, 'Starting applications', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 15, 14, 'Gaming', p=pattern_txt)
                    parseTxt(folder_name, tool, PCMARK7, 22, 21, 'Secondary storage score', p=pattern_score)
                     
                elif file == 'PCMARK8.xml':
                    PCMARK8 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----PCMARK8'
                    parseMark8Score(folder_name, tool, PCMARK8, storage_score)
                    parseMark8(folder_name, tool, PCMARK8, storage_bandwidth)					
				
                elif file == 'PCM_7.xml':
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----PCMARK8'
                    parseXmlGneric(folder_name, tool, current_dir_path, 'results/result')
                elif file == 'PCM_8.xml':
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----PCMARK8'
                    parseXmlGneric(folder_name, tool, current_dir_path, 'results/result')     
                elif file == 'SYSMARK2014.fdr':
                    SYSMARK2014 = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseFdr'
                    parseFdr(folder_name, tool, SYSMARK2014, overall_tag, scenario_tag, overallperformance_tag)
                     
                elif file == 'POR.txt':
                    POR = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]                        
                    print '----parseTxtPor/Spor'
                    parseTxtPor(folder_name, tool, POR, pattern_tBOOTUP, spec='POR time')
                     
                elif file == 'SPOR.txt':
                    SPOR = current_dir_path
                    folder_name = components[-3]
                    tool = components[-2]
                    print '----parseTxtPor/Spor'
                    parseTxtPor(folder_name, tool, SPOR, pattern_tBOOTUP, spec='SPOR time')
         
    con.commit()
    # file move to parsed
    for folders in gFolderInfos:
        src = pwd + '/' + folders['folderName']
        dst = pwd + '/archive/' + folders['folderName']
        if os.path.exists(dst):   
            shutil.rmtree(dst)
        shutil.copytree(src,dst)
        shutil.rmtree(src)
except MyError:
    print("test end")   
except cx_Oracle.DatabaseError ,ex:
    error, = ex.args    
    print query
    print 'Error Inserting Field Base'
    print 'Error.code    =', error.code
    print 'Error.message =', error.message
    print 'Error.offset  =', error.offset
    con.rollback()
    
finally :
    if cur:
        cur.close()     
    if con:
        con.close()   
        
    print "Complete" 